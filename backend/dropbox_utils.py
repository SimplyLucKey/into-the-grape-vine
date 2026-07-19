"""Shared Dropbox client and workbook helpers."""

from __future__ import annotations

import io
import logging
import time
from collections.abc import Callable
from functools import wraps
from typing import TypeVar

import dropbox
import dropbox.exceptions
import dropbox.files
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from dropbox_auth import get_dropbox_client

logger = logging.getLogger(__name__)

T = TypeVar("T")


def retry_dropbox(
    max_retries: int = 5,
) -> Callable[[Callable[..., T]], Callable[..., T]]:
    """Retry decorator for Dropbox operations with exponential backoff."""

    def decorator(func: Callable[..., T]) -> Callable[..., T]:
        @wraps(func)
        def wrapper(*args, **kwargs) -> T:  # type: ignore[no-untyped-def]
            for attempt in range(max_retries):
                try:
                    logger.info(
                        "%s attempt %d/%d... (Press Ctrl+C to abort)",
                        func.__name__,
                        attempt + 1,
                        max_retries,
                    )
                    return func(*args, **kwargs)
                except KeyboardInterrupt:
                    logger.warning("⚠ %s interrupted by user (Ctrl+C)", func.__name__)
                    raise
                except dropbox.exceptions.RateLimitError as e:
                    dropbox_retry_after = getattr(e, "retry_after", None)
                    retry_after = (
                        dropbox_retry_after if dropbox_retry_after else 5 * (2**attempt)
                    )

                    logger.error("=" * 80)
                    logger.error(
                        "✗ RATE LIMIT ERROR (attempt %d/%d)", attempt + 1, max_retries
                    )
                    logger.error("   Retry after: %s seconds", retry_after)

                    if attempt < max_retries - 1:
                        logger.warning(
                            "   → Waiting %ds before retry (Press Ctrl+C to abort)...",
                            retry_after,
                        )
                        try:
                            time.sleep(retry_after)
                        except KeyboardInterrupt:
                            logger.warning("⚠ Retry aborted by user")
                            raise
                    else:
                        logger.error("=" * 80)
                        logger.error(
                            "✗✗✗ FAILED: Rate limited after %d attempts", max_retries
                        )
                        logger.error("=" * 80)
                        raise
                except dropbox.exceptions.ApiError as e:
                    error_msg = str(e.error) if hasattr(e, "error") else str(e)
                    logger.error(
                        "✗ Dropbox API error (attempt %d/%d): %s",
                        attempt + 1,
                        max_retries,
                        error_msg,
                    )
                    if attempt < max_retries - 1:
                        wait_time = 2 ** (attempt + 1)
                        logger.warning(
                            "   → Retrying in %ds... (Press Ctrl+C to abort)", wait_time
                        )
                        try:
                            time.sleep(wait_time)
                        except KeyboardInterrupt:
                            logger.warning("⚠ Retry aborted by user")
                            raise
                    else:
                        raise
                except ConnectionError as e:
                    logger.error(
                        "✗ Connection error (attempt %d/%d): %s",
                        attempt + 1,
                        max_retries,
                        str(e),
                    )
                    if attempt < max_retries - 1:
                        wait_time = 2 ** (attempt + 1)
                        logger.warning(
                            "   → Retrying in %ds... (Press Ctrl+C to abort)", wait_time
                        )
                        try:
                            time.sleep(wait_time)
                        except KeyboardInterrupt:
                            logger.warning("⚠ Retry aborted by user")
                            raise
                    else:
                        raise
                except Exception as e:
                    logger.error(
                        "✗ Unexpected error (attempt %d/%d): %s",
                        attempt + 1,
                        max_retries,
                        type(e).__name__,
                    )
                    logger.error("   Details: %s", str(e))
                    if attempt < max_retries - 1:
                        wait_time = 2 ** (attempt + 1)
                        logger.warning(
                            "   → Retrying in %ds... (Press Ctrl+C to abort)", wait_time
                        )
                        try:
                            time.sleep(wait_time)
                        except KeyboardInterrupt:
                            logger.warning("⚠ Retry aborted by user")
                            raise
                    else:
                        raise
            raise RuntimeError("Unreachable")

        return wrapper

    return decorator


def get_client() -> dropbox.Dropbox:
    """Return an authenticated Dropbox client.

    Thin wrapper around get_dropbox_client() so callers import from one place.
    """
    return get_dropbox_client()


def download_workbook(
    client: dropbox.Dropbox,
    file_path: str,
) -> Workbook:
    """Download an Excel file from Dropbox and return it as a Workbook."""
    logger.info("Downloading workbook from Dropbox: %s", file_path)
    _, response = client.files_download(path=file_path)
    size_mb = len(response.content) / (1024 * 1024)
    logger.info("✓ Downloaded %s (%.2f MB)", file_path, size_mb)
    return load_workbook(filename=io.BytesIO(response.content))


def upload_workbook(
    client: dropbox.Dropbox,
    workbook: Workbook,
    file_path: str,
) -> None:
    """Serialize and upload a workbook to Dropbox, overwriting the existing file."""
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    content = buffer.read()
    file_size_mb = len(content) / (1024 * 1024)

    logger.info("Preparing upload: %s (%.2f MB)", file_path, file_size_mb)

    client.files_upload(
        f=content,
        path=file_path,
        mode=dropbox.files.WriteMode.overwrite,
    )
    logger.info("✓ Successfully uploaded workbook to %s", file_path)


def preview_workbook(workbook: Workbook, max_rows: int = 5) -> None:
    """Log a preview of each sheet — header plus up to max_rows data rows.

    Args:
        workbook: The loaded openpyxl Workbook.
        max_rows: Maximum number of data rows to log per sheet.
    """
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        logger.info(
            "Sheet '%s' (%d rows × %d cols)",
            sheet_name,
            sheet.max_row,
            sheet.max_column,
        )
        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_index > max_rows:
                logger.info("  ... (%d more rows)", sheet.max_row - max_rows - 1)
                break
            label = "header" if row_index == 0 else f"row {row_index:>3}"
            logger.info("  %s: %s", label, row)
