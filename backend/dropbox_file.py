"""Read-only access to the Vine tracker spreadsheet in Dropbox."""

import io
import os

import dropbox
import dropbox.exceptions
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from dropbox_auth import get_dropbox_client

load_dotenv()


def get_file_path() -> str:
    """Return the Dropbox file path from the environment.

    Raises:
        EnvironmentError: If DROPBOX_FILE_PATH is not set.
    """
    file_path: str | None = os.getenv("DROPBOX_FILE_PATH")
    if not file_path:
        raise EnvironmentError(
            "DROPBOX_FILE_PATH is not set in .env\n"
            "Example: /Vine/vine_orders.xlsx"
        )
    return file_path


def download_workbook(
    client: dropbox.Dropbox,
    file_path: str,
) -> Workbook:
    """Download an Excel file from Dropbox and return it as a Workbook.

    Args:
        client: Authenticated Dropbox client.
        file_path: Absolute path to the file within the user's Dropbox.

    Raises:
        dropbox.exceptions.ApiError: If the file cannot be found or accessed.
    """
    _, response = client.files_download(path=file_path)
    return load_workbook(filename=io.BytesIO(response.content))


def print_workbook_preview(workbook: Workbook, max_rows: int = 5) -> None:
    """Print a preview of each sheet — header plus up to max_rows data rows.

    Args:
        workbook: The loaded openpyxl Workbook.
        max_rows: Maximum number of data rows to print per sheet.
    """
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        print(f"\nSheet: '{sheet_name}' ({sheet.max_row} rows × {sheet.max_column} cols)")
        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if row_index > max_rows:
                remaining = sheet.max_row - max_rows - 1
                print(f"  ... ({remaining} more rows)")
                break
            label = "  header:" if row_index == 0 else f"  row {row_index:>3}:"
            print(f"{label} {row}")


def main() -> None:
    """Connect to Dropbox, download the spreadsheet, and print a preview."""
    print("── Connecting to Dropbox ──")
    client: dropbox.Dropbox = get_dropbox_client()
    account = client.users_get_current_account()
    print(f"✓ Connected as: {account.name.display_name} ({account.email})")

    file_path: str = get_file_path()
    print(f"\n── Downloading {file_path} ──")

    try:
        workbook: Workbook = download_workbook(client=client, file_path=file_path)
        print("✓ Downloaded successfully")
        print_workbook_preview(workbook=workbook, max_rows=5)
    except dropbox.exceptions.ApiError as error:
        if "not_found" in str(error):
            print(
                f"✗ File not found at '{file_path}'\n"
                "  Check DROPBOX_FILE_PATH in .env — path is relative to your Dropbox root."
            )
        else:
            print(f"✗ Dropbox API error: {error}")


if __name__ == "__main__":
    main()
