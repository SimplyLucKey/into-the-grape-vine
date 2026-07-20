"""FastAPI server for browser extension to sync data to Dropbox.

Provides HTTP endpoints that the extension can call instead of
accessing Dropbox directly. This keeps credentials secure in Python.

Usage:
    uv run python server.py
"""

from __future__ import annotations

import logging
from datetime import datetime

import dropbox
from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel

from dropbox_upsert import (
    INVENTORY_SHEET,
    _COL_DELIVERED_DATE,
    _COL_NAME,
    _COL_ORDER_DATE,
    _COL_PRICE,
    _COL_URL,
    append_orders_to_sheet,
    build_new_rows,
    extract_asin,
    get_existing_asins,
)
from dropbox_utils import download_workbook, get_client, upload_workbook
from fetch_prices import fetch_multiple_prices

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

app = FastAPI(title="Into the Grape Vine API")

# Allow extension to call this server from browser
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # Extension can call from any page
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class VineOrder(BaseModel):
    """Vine order data from extension."""

    asin: str | None
    name: str | None
    url: str | None
    thumbnail: str | None
    fmv: float | None
    order_date: str | None
    order_timestamp: int | None
    order_id: str | None
    captured_at: str


class AccountOrder(BaseModel):
    """Account order data from extension."""

    asin: str | None
    name: str | None
    url: str | None
    order_id: str | None
    delivery_status: str
    delivery_date: str | None
    delivery_date_parsed: str | None
    captured_at: str


class SyncRequest(BaseModel):
    """Request body for sync endpoint."""

    account_orders: list[AccountOrder]


class SyncVineOrdersRequest(BaseModel):
    """Request body for sync-vine-orders endpoint."""

    vine_orders: list[VineOrder]


class SyncVineOrdersResponse(BaseModel):
    """Response from sync-vine-orders endpoint."""

    success: bool
    added: int
    skipped: int
    dry_run: bool = False
    new_items: list[dict[str, str]] = []


class SyncResponse(BaseModel):
    """Response from sync endpoint."""

    success: bool
    filled: int
    cancelled: int
    cancelled_items: list[dict[str, str]]
    dry_run: bool = False
    changes: list[dict[str, str]] = []


def parse_delivery_date(order: AccountOrder) -> str | None:
    """Parse delivery date from account order, returning M/D/YYYY string."""
    if not order.delivery_date_parsed:
        return None

    dt = datetime.fromisoformat(order.delivery_date_parsed.replace("Z", "+00:00"))
    return dt.strftime("%-m/%-d/%Y")


def sync_delivery_dates_to_sheet(
    sheet: Worksheet,
    account_orders: list[AccountOrder],
    dry_run: bool = False,
) -> tuple[int, int, list[dict[str, str]], list[dict[str, str]]]:
    """Update blank delivered_date cells based on account orders.

    Args:
        sheet: The openpyxl inventory worksheet.
        account_orders: List of account orders from extension.
        dry_run: If True, don't actually modify the sheet, just report what would change.

    Returns:
        Tuple of (filled_count, cancelled_count, cancelled_items, changes).
    """
    # Index by ASIN for fast lookup
    orders_by_asin: dict[str, AccountOrder] = {}
    for order in account_orders:
        if order.asin:
            orders_by_asin[order.asin] = order

    filled = 0
    cancelled = 0
    cancelled_items: list[dict[str, str]] = []
    changes: list[dict[str, str]] = []

    # Iterate through Excel rows (skip header at row 1)
    for row_idx in range(2, sheet.max_row + 1):
        delivered_cell = sheet.cell(row=row_idx, column=_COL_DELIVERED_DATE)
        if delivered_cell.value:
            # Already has a delivery date
            continue

        url = sheet.cell(row=row_idx, column=_COL_URL).value
        if not url or not isinstance(url, str):
            continue

        asin = extract_asin(url=url)
        if not asin or asin not in orders_by_asin:
            continue

        order = orders_by_asin[asin]
        name = sheet.cell(row=row_idx, column=_COL_NAME).value

        if order.delivery_status == "delivered":
            delivery_date = parse_delivery_date(order)
            if delivery_date:
                filled += 1
                changes.append(
                    {
                        "row": str(row_idx),
                        "asin": asin,
                        "action": f"fill delivery date: {delivery_date}",
                        "product": name or "Unknown",
                    }
                )

                if not dry_run:
                    sheet.cell(
                        row=row_idx, column=_COL_DELIVERED_DATE, value=delivery_date
                    )

                action = "Would fill" if dry_run else "Filled"
                product_name = (name[:40] + "...") if name and len(name) > 40 else name
                logger.info(
                    "  → Row %d | %s | %s | %s delivery date: %s",
                    row_idx,
                    asin,
                    product_name or "Unknown",
                    action,
                    delivery_date,
                )

        elif order.delivery_status == "cancelled":
            cancelled += 1
            cancelled_items.append({"asin": asin, "name": name or "Unknown"})
            changes.append(
                {
                    "row": str(row_idx),
                    "asin": asin,
                    "action": "mark as cancelled (manual deletion recommended)",
                    "product": name or "Unknown",
                }
            )
            log_msg = f"Row {row_idx} ({asin}): Order cancelled"
            logger.warning(log_msg)

    return filled, cancelled, cancelled_items, changes


@app.post("/sync-vine-orders", response_model=SyncVineOrdersResponse)
async def sync_vine_orders(
    request: SyncVineOrdersRequest, dry_run: bool = False
) -> SyncVineOrdersResponse:
    """Sync Vine orders to Dropbox Excel file.

    This endpoint:
    1. Downloads the Excel file from Dropbox
    2. Checks which ASINs are already present
    3. Appends new Vine orders (order_date, url, name, FMV)
    4. Uploads the updated file back to Dropbox (unless dry_run=true)

    Args:
        request: Contains list of Vine orders to sync
        dry_run: If True, only report what would change without modifying the file
    """
    try:
        logger.info(
            "Received %d Vine orders to sync (dry_run=%s)",
            len(request.vine_orders),
            dry_run,
        )

        # Get Dropbox client
        client: dropbox.Dropbox = get_client()
        account = client.users_get_current_account()
        logger.info("Connected to Dropbox as %s", account.name.display_name)

        # Download workbook
        import os

        file_path: str | None = os.getenv("DROPBOX_FILE_PATH")
        if not file_path:
            raise HTTPException(
                status_code=500, detail="DROPBOX_FILE_PATH not configured"
            )

        workbook = download_workbook(client=client, file_path=file_path)

        if INVENTORY_SHEET not in workbook.sheetnames:
            raise HTTPException(
                status_code=500,
                detail=f"Sheet '{INVENTORY_SHEET}' not found in workbook",
            )

        sheet: Worksheet = workbook[INVENTORY_SHEET]

        # Get existing ASINs
        existing_asins = get_existing_asins(sheet=sheet)
        logger.info("Found %d existing ASINs in inventory", len(existing_asins))

        # Convert Pydantic models to dicts for dropbox_upsert functions
        orders_dicts = [order.model_dump() for order in request.vine_orders]

        # Build new rows (filters out existing ASINs)
        new_rows = build_new_rows(orders=orders_dicts, existing_asins=existing_asins)
        skipped = len(orders_dicts) - len(new_rows)

        logger.info(
            "%d new orders to insert, %d already present (skipped)",
            len(new_rows),
            skipped,
        )

        # Log each new order to be added
        new_items = []
        for order in new_rows:
            asin = order.get("asin", "")
            name = order.get("name", "Unknown")
            order_date = order.get("order_date", "")
            fmv = order.get("fmv")

            logger.info(
                "  → Adding: %s | %s | FMV: $%.2f | Date: %s",
                asin,
                name[:50] + "..." if len(name) > 50 else name,
                fmv if fmv else 0,
                order_date,
            )

            new_items.append(
                {
                    "asin": asin,
                    "name": name,
                    "order_date": order_date,
                }
            )

        if len(new_rows) == 0:
            logger.info("No new orders to add")
            return SyncVineOrdersResponse(
                success=True,
                added=0,
                skipped=skipped,
                dry_run=dry_run,
                new_items=[],
            )

        # Append orders (unless dry run)
        if not dry_run:
            logger.info("Adding %d orders to spreadsheet...", len(new_rows))
            added = append_orders_to_sheet(sheet=sheet, orders=new_rows)
            logger.info("Uploading modified spreadsheet to Dropbox...")
            upload_workbook(client=client, workbook=workbook, file_path=file_path)
            logger.info("✓ Vine sync complete: %d orders added to inventory", added)
        else:
            added = len(new_rows)
            logger.info("✓ DRY RUN: Would add %d orders to inventory", added)

        return SyncVineOrdersResponse(
            success=True,
            added=added,
            skipped=skipped,
            dry_run=dry_run,
            new_items=new_items,
        )

    except Exception as e:
        logger.exception("Vine order sync failed")
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/sync-delivery-dates", response_model=SyncResponse)
async def sync_delivery_dates(
    request: SyncRequest, dry_run: bool = False
) -> SyncResponse:
    """Sync delivery dates from account orders to Dropbox Excel file.

    This endpoint:
    1. Downloads the Excel file from Dropbox
    2. Matches account orders by ASIN
    3. Fills blank delivered_date cells for delivered items
    4. Uploads the updated file back to Dropbox (unless dry_run=true)

    Args:
        request: Contains list of account orders to sync
        dry_run: If True, only report what would change without modifying the file
    """
    try:
        logger.info(
            "Received %d account orders to sync (dry_run=%s)",
            len(request.account_orders),
            dry_run,
        )

        # Get Dropbox client
        client: dropbox.Dropbox = get_client()
        account = client.users_get_current_account()
        logger.info("Connected to Dropbox as %s", account.name.display_name)

        # Download workbook
        import os

        file_path: str | None = os.getenv("DROPBOX_FILE_PATH")
        if not file_path:
            raise HTTPException(
                status_code=500, detail="DROPBOX_FILE_PATH not configured"
            )

        workbook = download_workbook(client=client, file_path=file_path)

        if INVENTORY_SHEET not in workbook.sheetnames:
            raise HTTPException(
                status_code=500,
                detail=f"Sheet '{INVENTORY_SHEET}' not found in workbook",
            )

        sheet: Worksheet = workbook[INVENTORY_SHEET]

        # Sync delivery dates
        filled, cancelled, cancelled_items, changes = sync_delivery_dates_to_sheet(
            sheet=sheet,
            account_orders=request.account_orders,
            dry_run=dry_run,
        )

        if filled == 0 and cancelled == 0:
            logger.info("No updates needed")
            return SyncResponse(
                success=True,
                filled=0,
                cancelled=0,
                cancelled_items=[],
                dry_run=dry_run,
                changes=[],
            )

        # Upload updated workbook (unless dry run)
        if not dry_run:
            upload_workbook(client=client, workbook=workbook, file_path=file_path)
            logger.info("Sync complete: %d filled, %d cancelled", filled, cancelled)
        else:
            logger.info(
                "DRY RUN: Would fill %d, would mark %d as cancelled", filled, cancelled
            )

        return SyncResponse(
            success=True,
            filled=filled,
            cancelled=cancelled,
            cancelled_items=cancelled_items,
            dry_run=dry_run,
            changes=changes,
        )

    except Exception as e:
        logger.exception("Sync failed")
        raise HTTPException(status_code=500, detail=str(e))


class FetchProductPricesResponse(BaseModel):
    """Response from fetch-product-prices endpoint."""

    success: bool
    fetched: int
    failed: int
    skipped: int
    dry_run: bool = False


@app.post("/fetch-product-prices", response_model=FetchProductPricesResponse)
async def fetch_product_prices(
    dry_run: bool = False,
    days_back: int = 14,
    max_items: int = 50,
) -> FetchProductPricesResponse:
    """Fetch product prices from Amazon product pages for items missing prices.

    This endpoint:
    1. Downloads the Excel file from Dropbox
    2. Finds all rows with blank price column (within date threshold)
    3. Fetches current price from Amazon product pages
    4. Updates the price column (or marks as -1 if failed)
    5. Uploads the updated file back to Dropbox (unless dry_run=true)

    Args:
        dry_run: If True, only report what would be fetched without modifying the file
        days_back: Only fetch prices for orders within this many days (default 14)
        max_items: Maximum number of items to fetch in one run (default 50)
    """
    try:
        import os
        from datetime import datetime, timedelta

        logger.info(
            "Starting product price fetch (dry_run=%s, days_back=%d, max_items=%d)",
            dry_run,
            days_back,
            max_items,
        )

        # Get Dropbox client
        client: dropbox.Dropbox = get_client()
        account = client.users_get_current_account()
        logger.info("Connected to Dropbox as %s", account.name.display_name)

        # Download workbook
        file_path: str | None = os.getenv("DROPBOX_FILE_PATH")
        if not file_path:
            raise HTTPException(
                status_code=500, detail="DROPBOX_FILE_PATH not configured"
            )

        workbook = download_workbook(client=client, file_path=file_path)

        if INVENTORY_SHEET not in workbook.sheetnames:
            raise HTTPException(
                status_code=500,
                detail=f"Sheet '{INVENTORY_SHEET}' not found in workbook",
            )

        sheet: Worksheet = workbook[INVENTORY_SHEET]

        # Calculate cutoff date
        cutoff_date = datetime.now() - timedelta(days=days_back)
        logger.info(
            "Only fetching product prices for orders after %s",
            cutoff_date.strftime("%Y-%m-%d"),
        )

        # Find rows with missing product prices (within date threshold)
        asins_to_fetch: list[tuple[int, str]] = []  # (row_idx, asin)

        for row_idx in range(2, sheet.max_row + 1):
            price_cell = sheet.cell(row=row_idx, column=_COL_PRICE)

            # Skip if already has a price (including -1 for failed attempts)
            if price_cell.value is not None:
                continue

            # Check order date
            order_date_cell = sheet.cell(row=row_idx, column=_COL_ORDER_DATE)
            if order_date_cell.value:
                try:
                    if isinstance(order_date_cell.value, datetime):
                        order_date = order_date_cell.value
                    else:
                        # Try parsing if it's a string
                        order_date = datetime.strptime(
                            str(order_date_cell.value), "%m/%d/%Y"
                        )

                    # Skip if too old
                    if order_date < cutoff_date:
                        continue
                except (ValueError, AttributeError):
                    # Can't parse date, skip this row
                    logger.warning("Row %d: Could not parse order date", row_idx)
                    continue

            url = sheet.cell(row=row_idx, column=_COL_URL).value
            if not url or not isinstance(url, str):
                continue

            asin = extract_asin(url=url)
            if asin:
                asins_to_fetch.append((row_idx, asin))

                # Stop at max_items limit
                if len(asins_to_fetch) >= max_items:
                    logger.info(
                        "Reached max_items limit (%d), stopping scan", max_items
                    )
                    break

        if not asins_to_fetch:
            logger.info("No items need product price fetching")
            return FetchProductPricesResponse(
                success=True,
                fetched=0,
                failed=0,
                skipped=0,
                dry_run=dry_run,
            )

        logger.info("Found %d items needing product prices", len(asins_to_fetch))

        if dry_run:
            logger.info(
                "DRY RUN: Would fetch product prices for %d items", len(asins_to_fetch)
            )
            return FetchProductPricesResponse(
                success=True,
                fetched=0,
                failed=0,
                skipped=len(asins_to_fetch),
                dry_run=True,
            )

        # Fetch product prices from Amazon
        asins_only = [asin for _, asin in asins_to_fetch]
        prices = fetch_multiple_prices(asins_only)

        # Update sheet
        fetched = 0
        failed = 0

        for row_idx, asin in asins_to_fetch:
            price = prices.get(asin)
            if price is not None:
                sheet.cell(row=row_idx, column=_COL_PRICE, value=price)
                fetched += 1
                logger.info(
                    "Row %d (%s): Set product price $%.2f", row_idx, asin, price
                )
            else:
                # Mark as -1 to indicate we tried and failed
                # This prevents retrying the same item every time
                sheet.cell(row=row_idx, column=_COL_PRICE, value=-1)
                failed += 1
                logger.warning(
                    "Row %d (%s): Failed to fetch product price, marked as -1",
                    row_idx,
                    asin,
                )

        # Upload updated workbook (even if some failed, to save -1 markers)
        if fetched > 0 or failed > 0:
            upload_workbook(client=client, workbook=workbook, file_path=file_path)
            logger.info(
                "Product price fetch complete: %d fetched, %d failed", fetched, failed
            )

        return FetchProductPricesResponse(
            success=True,
            fetched=fetched,
            failed=failed,
            skipped=0,
            dry_run=False,
        )

    except Exception as e:
        logger.exception("Product price fetch failed")
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/health")
async def health_check() -> dict[str, str]:
    """Health check endpoint."""
    return {"status": "ok"}


if __name__ == "__main__":
    import uvicorn

    logger.info("Starting server on http://localhost:8000")
    uvicorn.run(app, host="127.0.0.1", port=8000)
