"""Upserts new Vine orders into the inventory sheet of the Dropbox Excel file.

Upsert key: ASIN extracted from the product URL.
- Existing ASIN: skip entirely — never overwrite any fields.
- New ASIN: append in ascending order_date order (oldest first).

Columns written on insert:
    order_date, url, name, FMV

Pending future implementation:
    delivered_date  — requires order detail page visit
    price           — requires product listing page visit
    rating          — populated by LLM review pipeline
    reviewed        — populated by LLM review pipeline
    verbatim        — populated by LLM review pipeline

Out of scope:
    sold, sell_price
"""

from __future__ import annotations

import logging
import os
import re
from datetime import datetime
from typing import Any

import dropbox
from dotenv import load_dotenv
from openpyxl.worksheet.worksheet import Worksheet

from dropbox_utils import download_workbook, get_client, upload_workbook

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

_ASIN_PATTERN: re.Pattern[str] = re.compile(r"/dp/([A-Z0-9]{10})")

INVENTORY_SHEET: str = "inventory"

# Column positions (1-based) matching the inventory sheet header row
_COL_ORDER_DATE: int = 1
_COL_DELIVERED_DATE: int = 2
_COL_URL: int = 3
_COL_NAME: int = 4
_COL_FMV: int = 5
_COL_PRICE: int = 6
_COL_RATING: int = 7
_COL_REVIEWED: int = 8
_COL_VERBATIM: int = 9
_COL_SOLD: int = 10
_COL_SELL_PRICE: int = 11


def extract_asin(url: str) -> str | None:
    """Extract the ASIN from an Amazon product URL.

    Args:
        url: Amazon product URL containing /dp/XXXXXXXXXX/.

    Returns:
        The 10-character ASIN string, or None if not found.
    """
    match = _ASIN_PATTERN.search(url)
    return match.group(1) if match else None


def get_existing_asins(sheet: Worksheet) -> set[str]:
    """Return ASINs already present in the inventory sheet.

    Reads the URL column (col C) for every data row.

    Args:
        sheet: The openpyxl inventory worksheet.

    Returns:
        Set of ASIN strings currently in the sheet.
    """
    asins: set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        url = row[_COL_URL - 1]
        if url and isinstance(url, str):
            asin = extract_asin(url=url)
            if asin:
                asins.add(asin)
    return asins


def parse_order_date(order: dict[str, Any]) -> str | None:
    """Parse order date, returning YYYY-MM-DD string. Prefers order_timestamp over order_date."""
    if order.get("order_timestamp"):
        dt = datetime.fromtimestamp(order["order_timestamp"] / 1000)
        return dt.strftime("%Y-%m-%d")

    date_str: str | None = order.get("order_date")
    if not date_str:
        return None

    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%B %d, %Y"):
        try:
            dt = datetime.strptime(date_str, fmt)
            return dt.strftime("%Y-%m-%d")
        except ValueError:
            continue

    logger.warning("Could not parse order_date '%s'", date_str)
    return None


def build_new_rows(
    orders: list[dict[str, Any]],
    existing_asins: set[str],
) -> list[dict[str, Any]]:
    """Filter and sort orders not already in the inventory sheet.

    Args:
        orders: List of order dicts from the extension queue.
        existing_asins: ASINs already present in the sheet.

    Returns:
        New orders sorted oldest-first, ready to append.
    """
    new_orders: list[dict[str, Any]] = [
        order
        for order in orders
        if order.get("url")
        and (asin := extract_asin(url=order["url"]))
        and asin not in existing_asins
    ]
    new_orders.sort(key=lambda o: parse_order_date(order=o) or datetime.min)
    return new_orders


def append_orders_to_sheet(
    sheet: Worksheet,
    orders: list[dict[str, Any]],
) -> int:
    """Append new orders as rows to the inventory sheet.

    Args:
        sheet: The openpyxl inventory worksheet.
        orders: New orders to append, sorted oldest-first.

    Returns:
        Number of rows appended.
    """
    # Find the actual last row with data (not just max_row which can include empty rows)
    last_data_row = 1  # Start at header row
    for row_idx in range(2, sheet.max_row + 1):
        # Check if this row has any data in key columns
        if (
            sheet.cell(row=row_idx, column=_COL_ORDER_DATE).value
            or sheet.cell(row=row_idx, column=_COL_URL).value
            or sheet.cell(row=row_idx, column=_COL_NAME).value
        ):
            last_data_row = row_idx

    next_row: int = last_data_row + 1

    for order in orders:
        sheet.cell(
            row=next_row, column=_COL_ORDER_DATE, value=parse_order_date(order=order)
        )
        sheet.cell(row=next_row, column=_COL_URL, value=order.get("url"))
        sheet.cell(row=next_row, column=_COL_NAME, value=order.get("name"))
        sheet.cell(row=next_row, column=_COL_FMV, value=order.get("fmv"))
        next_row += 1

    return len(orders)


def upsert_orders(orders: list[dict[str, Any]]) -> None:
    """Upsert a list of orders into the Dropbox inventory sheet.

    Args:
        orders: List of order dicts from the extension queue. Each must contain
                at minimum: url, name, fmv, and order_date or order_timestamp.
    """
    file_path: str | None = os.getenv("DROPBOX_FILE_PATH")
    if not file_path:
        raise EnvironmentError("DROPBOX_FILE_PATH is not set in .env")

    client: dropbox.Dropbox = get_client()
    account = client.users_get_current_account()
    logger.info("Connected as %s", account.name.display_name)

    workbook = download_workbook(client=client, file_path=file_path)

    if INVENTORY_SHEET not in workbook.sheetnames:
        raise ValueError(
            f"Sheet '{INVENTORY_SHEET}' not found. Available: {workbook.sheetnames}"
        )

    sheet: Worksheet = workbook[INVENTORY_SHEET]
    existing_asins = get_existing_asins(sheet=sheet)
    logger.info("Found %d existing ASINs in inventory", len(existing_asins))

    new_rows = build_new_rows(orders=orders, existing_asins=existing_asins)
    skipped = len(orders) - len(new_rows)
    logger.info(
        "%d new orders to insert, %d already present (skipped)", len(new_rows), skipped
    )

    if not new_rows:
        logger.info("Nothing to do.")
        return

    appended = append_orders_to_sheet(sheet=sheet, orders=new_rows)
    upload_workbook(client=client, workbook=workbook, file_path=file_path)
    logger.info("Uploaded — %d rows added to '%s'", appended, INVENTORY_SHEET)


if __name__ == "__main__":
    # Smoke test — inserts one fake row to verify the pipeline end-to-end.
    # Delete the inserted row from the sheet after confirming it worked.
    _test_orders: list[dict[str, Any]] = [
        {
            "asin": "B0TEST00001",
            "name": "Test Product — delete this row",
            "url": "https://www.amazon.com/dp/B0TEST00001",
            "fmv": 9.99,
            "order_date": "1/1/2025",
            "order_timestamp": None,
        }
    ]
    upsert_orders(orders=_test_orders)
